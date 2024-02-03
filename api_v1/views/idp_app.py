import pendulum
from django.contrib.auth import get_user_model
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from drf_spectacular.utils import extend_schema
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.filters import OrderingFilter, SearchFilter
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.request import Request
from rest_framework.response import Response

from api_v1.serializers.idp_app import (
    CreateIDPSerializer,
    DepartmentSerializer,
    FileSerializer,
    IDPasFieldSerializer,
    IDPNotificationSerializer,
    IDPReadOnlySerializer,
    NotificationSerializer,
    TaskNotificationSerializer,
    TaskSerializer,
)
from core.choices import IdpStatuses, StatusChoices
from core.constants import DEFAULT_COLUMN_WIDTH, IDP_NAME_COLUMN_WIDTH
from core.utils import get_idp_extra_info
from idp_app.models import (
    IDP,
    File,
    IdpNotification,
    Notification,
    Task,
    TaskNotification,
)
from users.models import Department

User = get_user_model()


class DepartmentViewSet(viewsets.ModelViewSet):
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer
    permission_classes = (AllowAny,)


class TaskViewSet(viewsets.ModelViewSet):
    queryset = Task.objects.all()
    serializer_class = TaskSerializer
    permission_classes = (AllowAny,)


class FileViewSet(viewsets.ModelViewSet):
    queryset = File.objects.all()
    serializer_class = FileSerializer
    permission_classes = (AllowAny,)

    def download_file(self, request, file_id):
        file_obj = get_object_or_404(File, pk=file_id)
        file_content = file_obj.file.read()
        response = HttpResponse(
            file_content, content_type="application/octet-stream"
        )
        response[
            "Content-Disposition"
        ] = f'attachment; filename="{file_obj.file.name}"'
        return response


class NotificationViewSet(viewsets.ModelViewSet):
    queryset = Notification.objects.all()
    serializer_class = NotificationSerializer
    permission_classes = (AllowAny,)


class IDPViewSet(viewsets.ModelViewSet):
    queryset = IDP.objects.all()
    serializer_class = IDPReadOnlySerializer
    permission_classes = (IsAuthenticated,)
    filter_backends = (SearchFilter, OrderingFilter)
    search_fields = ("^name", "employee__last_name")
    ordering_fields = (
        "name",
        "end_date_plan",
        "status",
        "employee__last_name",
        "employee__first_name",
    )

    def get_serializer_class(self):
        if self.request.method == "GET":
            return IDPReadOnlySerializer
        return CreateIDPSerializer

    def create(self, request, *args, **kwargs):
        user = request.user
        if user.is_authenticated:
            idp_data = request.data

            employee_cheif = User.objects.get(uid=idp_data["employee"]).chief

            if user == employee_cheif:
                idp_data["status"] = StatusChoices.ACTIVE
            else:
                idp_data["status"] = StatusChoices.DRAFT

            serializer = CreateIDPSerializer(data=idp_data)
            serializer.is_valid(raise_exception=True)
            serializer.save()

            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(status=status.HTTP_401_UNAUTHORIZED)

    @extend_schema(responses=IDPasFieldSerializer(many=True))
    @action(detail=False, url_path="private")
    def get_private_idps(self, request: Request):
        """Возвращает список личных ипр авторизованного пользователя."""
        idps = request.user.idps.all().order_by("-end_date_plan")
        filtered_idps = self.filter_queryset(idps)
        if filtered_idps:
            extra_info = get_idp_extra_info(filtered_idps)
            page = self.paginate_queryset(filtered_idps)
            if page:
                serializer = IDPasFieldSerializer(page, many=True)
                response = self.get_paginated_response(serializer.data)
            else:
                serializer = IDPasFieldSerializer(filtered_idps, many=True)
                response = Response(data=serializer.data)

            extra_info.update(response.data)
            response.data = extra_info
            return response
        return Response({"detail": "У вас еще нет ИПР."})

    @extend_schema(responses=IDPasFieldSerializer(many=True))
    @action(detail=False, url_path="subordinates")
    def get_subordinates_idps(self, request: Request):
        """Возвращает список ипр подчиненных."""
        subordinates = request.user.subordinates.all()
        if not subordinates:
            return Response({"detail": "У вас нет сотрудников."})
        idps = (
            IDP.objects.filter(employee__in=subordinates)
            .exclude(status=IdpStatuses.DRAFT)
            .order_by("-end_date_plan", "employee__last_name")
        )
        filtered_idps = self.filter_queryset(idps)

        if filtered_idps:
            extra_info = get_idp_extra_info(filtered_idps)
            page = self.paginate_queryset(filtered_idps)
            if page:
                serializer = IDPasFieldSerializer(page, many=True)
                response = self.get_paginated_response(serializer.data)
            else:
                serializer = IDPasFieldSerializer(filtered_idps, many=True)
                response = Response(data=serializer.data)

            extra_info.update(response.data)
            response.data = extra_info

            return response

        return Response({"detail": "У ваших сотрудников еще нет ИПР."})

    @action(detail=False, url_path="export/excel")
    def export_idps_to_excel(self, request):
        """Экспорт списка ИПР подчиненных в excel файл."""
        subordinates_data = self.get_subordinates_idps(request).data

        if subordinates_data:
            workbook = Workbook()
            excel_sheet = workbook.active
            headers = [
                "План развития",
                "Cотрудник",
                "Плановая дата закрытия",
                "Фактическая дата закрытия",
                "Статус",
            ]
            for col_num, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_num)
                cell = excel_sheet[f"{col_letter}1"]
                cell.value = header
                cell.font = Font(bold=True)
                # Устанавливаем ширину столбцов по умолчанию
                excel_sheet.column_dimensions[
                    col_letter
                ].width = DEFAULT_COLUMN_WIDTH
                excel_sheet.column_dimensions[
                    "A"
                ].width = IDP_NAME_COLUMN_WIDTH
            for row_num, idp in enumerate(subordinates_data["results"], 2):
                excel_sheet[f"A{row_num}"] = idp.get("name", "")
                employee = idp.get("employee", {})
                employee_name = f"{employee.get('first_name', '')} {employee.get('last_name', '')}"
                excel_sheet[f"B{row_num}"] = employee_name
                time_str = idp.get("end_date_plan", "")
                # Дату приводим к формату DD:MM:YYYY HH:mm
                end_date_plan = pendulum.parse(time_str)
                formatted_date = end_date_plan.format("DD:MM:YYYY HH:mm")
                excel_sheet[f"C{row_num}"] = formatted_date
                excel_sheet[f"D{row_num}"] = idp.get("end_date_fact", "")
                status_eng = idp.get("status", "")
                status_label = getattr(IdpStatuses, status_eng.upper()).label
                excel_sheet[f"E{row_num}"] = status_label
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response[
                "Content-Disposition"
            ] = "attachment; filename=subordinates_idps.xlsx"
            workbook.save(response)

            return response

        return Response({"detail": "Нет данных для экспорта в Excel."})


class TaskNotificationViewSet(viewsets.ModelViewSet):
    queryset = TaskNotification.objects.all()
    serializer_class = TaskNotificationSerializer
    permission_classes = (AllowAny,)


class IDPNotificationViewSet(viewsets.ModelViewSet):
    queryset = IdpNotification.objects.all()
    serializer_class = IDPNotificationSerializer
    permission_classes = (AllowAny,)
