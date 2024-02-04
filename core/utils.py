from datetime import timedelta
from typing import Dict, List

from django.db.models import Case, Model, Value, When
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pendulum import parse as pendulum_parse

from core.choices import IdpStatuses, TaskStatuses
from core.constants import DEFAULT_COLUMN_WIDTH, IDP_NAME_COLUMN_WIDTH


def default_end_date_plan():
    return timezone.now() + timedelta(days=30)


def find_differencies(initial: Model, updated: Model):
    """
    Возвращает различия между полями двух объектов.

    Сравнивает два dict исходного и обновленного объектов.
    Если какое-либо поле отличается, мы берем обновленное значение.
    Возвращает dict с новыми значениями.
    """
    d1 = initial.__dict__
    d2 = updated.__dict__
    diffs = {k: v for k, v in d2.items() if v != d1[k]}
    if "_state" in diffs:
        # _state нам без надобности, поэтому убираем
        del diffs["_state"]
    return diffs


def get_idp_extra_info(idps: List[Model]) -> Dict[str, int]:
    """
    Генерирует дополнительные поля для ответа на запрос ИПР.

    Считает количество ипр по статусам.
    """
    extra_info = {
        "in_total": 0,
        IdpStatuses.ACTIVE: 0,
        IdpStatuses.CLOSED: 0,
        IdpStatuses.OVERDUE: 0,
    }
    for idp in idps:
        if idp.status in (
            IdpStatuses.ACTIVE,
            IdpStatuses.TWO_WEEKS,
            IdpStatuses.COMPLETED_APPROVAL,
            IdpStatuses.DRAFT_APPROVAL,
        ):
            extra_info["in_total"] += 1
            extra_info[IdpStatuses.ACTIVE] += 1
        elif idp.status == IdpStatuses.CLOSED:
            extra_info["in_total"] += 1
            extra_info[IdpStatuses.CLOSED] += 1
        elif idp.status == IdpStatuses.OVERDUE:
            extra_info["in_total"] += 1
            extra_info[IdpStatuses.OVERDUE] += 1
    return extra_info


def get_extensions():
    """Возвращает списки допустимых разрешений и MIME типов для загрузки файлов."""
    extension_mapping = {
        ".pdf": "application/pdf",
        ".doc": "application/msword",
        ".xls": "application/vnd.ms-excel",
        ".jpeg": "image/jpeg",
        ".jpg": "image/jpeg",
        ".png": "image/png",
        ".zip": "application/zip",
        ".rar": "application/x-rar-compressed",
    }
    extensions = list(extension_mapping.keys())
    content_types = list(extension_mapping.values())
    return extensions, content_types


def setup_excel_file(data):
    """Настройка Excel-файла для экспорта ИПР подчиненных."""
    workbook = Workbook()
    excel_sheet = workbook.active
    headers = [
        "План развития",
        "Cотрудник",
        "Плановая дата закрытия",
        "Фактическая дата закрытия",
        "Статус",
    ]
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = excel_sheet[f"{col_letter}1"]
        cell.value = header
        cell.font = Font(bold=True)
        excel_sheet.column_dimensions[col_letter].width = DEFAULT_COLUMN_WIDTH
        excel_sheet.column_dimensions["A"].width = IDP_NAME_COLUMN_WIDTH
    for row_num, idp in enumerate(data["results"], 2):
        excel_sheet[f"A{row_num}"] = idp.get("name", "")
        employee = idp.get("employee", {})
        employee_name = (
            f"{employee.get('first_name', '')} {employee.get('last_name', '')}"
        )
        excel_sheet[f"B{row_num}"] = employee_name
        time_str = idp.get("end_date_plan", "")
        end_date_plan = pendulum_parse(time_str)
        formatted_date = end_date_plan.format("DD:MM:YYYY HH:mm")
        excel_sheet[f"C{row_num}"] = formatted_date
        excel_sheet[f"D{row_num}"] = idp.get("end_date_fact", "")
        status_eng = idp.get("status", "")
        status_label = getattr(IdpStatuses, status_eng.upper()).label
        excel_sheet[f"E{row_num}"] = status_label
    return workbook
  
  
# Упорядочивает по статусам ипр
idp_status_order = Case(
    When(status=IdpStatuses.DRAFT_APPROVAL, then=Value(1)),
    When(status=IdpStatuses.ACTIVE, then=Value(2)),
    When(status=IdpStatuses.TWO_WEEKS, then=Value(3)),
    When(status=IdpStatuses.OVERDUE, then=Value(4)),
    When(status=IdpStatuses.COMPLETED_APPROVAL, then=Value(5)),
    When(status=IdpStatuses.CLOSED, then=Value(6)),
    When(status=IdpStatuses.CANCELLED, then=Value(7)),
)

# Упорядочивает по статусам задач
task_status_order = Case(
    When(task_status=TaskStatuses.DRAFT_APPROVAL, then=Value(1)),
    When(task_status=TaskStatuses.ACTIVE, then=Value(2)),
    When(task_status=TaskStatuses.TWO_WEEKS, then=Value(3)),
    When(task_status=TaskStatuses.OVERDUE, then=Value(4)),
    When(task_status=TaskStatuses.COMPLETED_APPROVAL, then=Value(5)),
    When(task_status=TaskStatuses.CLOSED, then=Value(6)),
    When(task_status=TaskStatuses.CANCELLED, then=Value(6)),
)
